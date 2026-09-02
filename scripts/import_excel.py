#!/usr/bin/env python3
"""Import a monthly Work_Order_MM_YYYY.xlsx workbook into the staging schema.

Usage:
    python3 scripts/import_excel.py path/to/Work_Order_08_2026.xlsx

Reads columns A-H only (No., WO#, Date, Category, Location, Description,
Status, By); ignores the I-L pivot-formula columns, column O, and all cell
colours (§1: colour is conditional formatting derived from the category, not
independent data). Writes work orders and log entries into the `staging`
schema (run scripts/staging_schema.sql first) and upserts any new category /
location strings straight into the public reference tables. Produces an
import summary on stdout and a `review_needed.xlsx` file next to the input
workbook listing every row that needs a human look before promotion.

Re-runnable: truncates the staging tables at the start of each run.

See the build spec §6 for the full rule set this implements.
"""
from __future__ import annotations

import os
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import date

import openpyxl
import psycopg2
import psycopg2.extras

# ---------------------------------------------------------------------------
# Documented corrections (§6 step 12) — the ONLY deviations from the sheet.
# Keep this list exhaustive and visible; do not generalise either entry into
# a rule that could silently touch other rows.
# ---------------------------------------------------------------------------
CATEGORY_CORRECTIONS = {
    # wo_no -> (from, to, reason)
    "2608007": ("ล้างแอร์", "งานประจำ", "Miscategorised; confirmed not paid"),
}

# §6 step 5: WO# 2608131 is a hand-resolved exception, hard-coded rather than
# generalised — one job, category แอร์, all 7 rows become log entries.
HARDCODED_MERGE_WO = "2608131"
HARDCODED_MERGE_CATEGORY = "แอร์"
HARDCODED_MERGE_REASON = (
    "6 rows category แอร์ + 1 row (14 Aug) category เช็คแอร์ — same Lobby unit, "
    "same fault, imported as one job, category แอร์ (see spec §6 step 5/12)."
)

# §6 step 5: these two collisions touch special (paid) categories. A wrong
# auto-split would create or destroy a payable job, so they are NOT split by
# this script — every row is routed to review_needed.xlsx untouched.
MANUAL_REVIEW_ONLY_WO = {"2608025", "2608044"}

# §2 workers seed — used to split/match the free-text By column.
KNOWN_WORKERS = ["เปิ้ล", "นา", "ข้าง", "ฮอง", "ช่างมิตซูบิชิ", "OutSource", "จัดซื้อ", "รอสี"]

# §6 step 10: known near-duplicate location spellings, flagged (not merged).
NEAR_DUP_LOCATION_PAIRS = [
    ("ห้องน้ำคนขับ", "ห้องน้ำคนขับรถ"),
    ("Lobby", "ล็อบบี้"),
]

HEADER_MARKERS = {"WO#", "Work Orders#", "No."}

DONE = "เสร็จ"
PENDING = "ค้าง"


def nfc(s) -> str:
    return unicodedata.normalize("NFC", str(s).strip()) if s is not None else ""


@dataclass
class RawRow:
    excel_row: int
    wo_no: str
    date_raw: object
    category_raw: str
    location_raw: str
    description: str
    status_raw: str
    by_raw: str
    parsed_date: date | None = None
    status: str | None = None  # 'pending' | 'done'
    workers: list[str] = field(default_factory=list)
    unmatched_worker_text: str | None = None


@dataclass
class ReviewItem:
    severity: str  # 'must_resolve' | 'fyi'
    wo_no: str
    excel_row: object
    reason: str
    original_value: str = ""


def parse_date(raw) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", s)
    if not m:
        return None
    dd, mm, yyyy = (int(x) for x in m.groups())
    year = yyyy
    assert 2020 <= year <= 2030, (
        f"Parsed year {year} out of [2020, 2030] range for '{s}' — looks like a "
        "Buddhist-era date or bad cell; failing loudly rather than importing a "
        "date 543 years out (§6 step 6)."
    )
    return date(year, mm, dd)


def map_status(raw) -> str | None:
    v = nfc(raw)
    if v == DONE:
        return "done"
    if v == PENDING:
        return "pending"
    return None


def split_workers(raw) -> tuple[list[str], str | None]:
    """Split the By column on '/', ',', whitespace, then longest-match
    unseparated fragments (e.g. "นาข้าง") against known worker names.
    Returns (matched_names, unmatched_leftover_text_or_None)."""
    if not raw or not str(raw).strip():
        return [], None
    fragments = [f for f in re.split(r"[,/\s]+", nfc(raw)) if f]
    matched: list[str] = []
    unmatched_parts: list[str] = []
    names_by_len = sorted(KNOWN_WORKERS, key=len, reverse=True)
    for frag in fragments:
        if frag in KNOWN_WORKERS:
            matched.append(frag)
            continue
        # Greedy longest-match decomposition for no-separator fragments.
        remaining = frag
        local_matched: list[str] = []
        while remaining:
            for name in names_by_len:
                if remaining.startswith(name):
                    local_matched.append(name)
                    remaining = remaining[len(name):]
                    break
            else:
                break
        if remaining:
            unmatched_parts.append(frag)
        else:
            matched.extend(local_matched)
    return matched, (" ".join(unmatched_parts) or None)


def infer_location_type_and_floor(code: str) -> tuple[str, int | None]:
    if code == "คอนคอร์ด":
        return "external", None
    if re.fullmatch(r"\d{3,4}", code):
        floor = int(code[:-2]) if len(code) == 4 else int(code[0])
        return "room", floor
    return "common", None


class Importer:
    def __init__(self, conn):
        self.conn = conn
        self.review: list[ReviewItem] = []
        self.category_cache: dict[str, str] = {}
        self.location_cache: dict[str, str] = {}
        self.wo_seq_cache: dict[str, int] = {}
        self.imported_user_id = self._fetch_imported_user_id()

    # -- reference data (public schema): upsert-as-you-go -------------------
    def _fetch_imported_user_id(self) -> str:
        with self.conn.cursor() as cur:
            cur.execute("SELECT id FROM users WHERE username = 'imported'")
            row = cur.fetchone()
            if not row:
                raise RuntimeError(
                    "System user 'imported' not found — run `npm run db:seed` first."
                )
            return row[0]

    def get_or_create_category(self, name_th: str) -> str:
        if name_th in self.category_cache:
            return self.category_cache[name_th]
        with self.conn.cursor() as cur:
            cur.execute("SELECT id FROM categories WHERE name_th = %s", (name_th,))
            row = cur.fetchone()
            if not row:
                # Not one of the 22 seeded values — unexpected. Create it
                # conservatively (not special, no group) and flag for review
                # rather than silently deciding it's payable.
                cur.execute(
                    "INSERT INTO categories (name_th, is_special) VALUES (%s, false) RETURNING id",
                    (name_th,),
                )
                row = cur.fetchone()
                self.review.append(
                    ReviewItem(
                        "must_resolve", "", "", f"Category '{name_th}' was not in the seeded "
                        "22 — created as not-special. Confirm is_special/group in admin.",
                        name_th,
                    )
                )
            self.category_cache[name_th] = row[0]
            return row[0]

    def get_or_create_location(self, code: str) -> str:
        if code in self.location_cache:
            return self.location_cache[code]
        with self.conn.cursor() as cur:
            cur.execute("SELECT id FROM locations WHERE code = %s", (code,))
            row = cur.fetchone()
            if not row:
                loc_type, floor = infer_location_type_and_floor(code)
                cur.execute(
                    "INSERT INTO locations (code, type, floor) VALUES (%s, %s, %s) RETURNING id",
                    (code, loc_type, floor),
                )
                row = cur.fetchone()
            self.location_cache[code] = row[0]
        for a, b in NEAR_DUP_LOCATION_PAIRS:
            if code in (a, b):
                self.review.append(
                    ReviewItem(
                        "fyi", "", "", f"Location '{code}' is a near-duplicate spelling "
                        f"of '{b if code == a else a}' — both imported as separate "
                        "locations; deactivate the unwanted one in admin.",
                        code,
                    )
                )
        return self.location_cache[code]

    def get_worker_id(self, name: str) -> str | None:
        with self.conn.cursor() as cur:
            cur.execute("SELECT id FROM workers WHERE name = %s", (name,))
            row = cur.fetchone()
            return row[0] if row else None

    # -- WO# generation for split siblings -----------------------------------
    def next_wo_no(self, prefix: str) -> str:
        if prefix not in self.wo_seq_cache:
            with self.conn.cursor() as cur:
                cur.execute(
                    "SELECT COALESCE(MAX(CAST(RIGHT(wo_no, 3) AS INT)), 0) "
                    "FROM staging.work_orders WHERE wo_no LIKE %s",
                    (prefix + "%",),
                )
                self.wo_seq_cache[prefix] = cur.fetchone()[0]
        self.wo_seq_cache[prefix] += 1
        return f"{prefix}{self.wo_seq_cache[prefix]:03d}"

    # -- staging writes -------------------------------------------------------
    def insert_work_order(
        self, wo_no: str, legacy_wo_no: str | None, rows: list[RawRow], category_th: str, location_code: str
    ) -> int:
        dated = [r for r in rows if r.parsed_date is not None]
        if not dated:
            self.review.append(
                ReviewItem(
                    "must_resolve", wo_no, [r.excel_row for r in rows],
                    "No row in this work order has a parseable date (date cell "
                    "contains 'ค้าง' or similar) — cannot set opened_date. Not imported.",
                )
            )
            return 0
        dated.sort(key=lambda r: (r.parsed_date, r.excel_row))
        opened_date = dated[0].parsed_date
        final_status = dated[-1].status

        closed_date = None
        for r in dated:
            if r.status == "done":
                if closed_date is None:
                    closed_date = r.parsed_date
            else:
                closed_date = None
        if final_status != "done":
            closed_date = None

        category_id = self.get_or_create_category(category_th)
        location_id = self.get_or_create_location(location_code)

        with self.conn.cursor() as cur:
            cur.execute(
                """INSERT INTO staging.work_orders
                   (wo_no, legacy_wo_no, opened_date, category_id, location_id,
                    description, status, created_by, closed_date)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id""",
                (
                    wo_no, legacy_wo_no, opened_date, category_id, location_id,
                    dated[0].description, final_status, self.imported_user_id, closed_date,
                ),
            )
            work_order_id = cur.fetchone()[0]

            for r in dated:
                cur.execute(
                    """INSERT INTO staging.wo_log_entries
                       (work_order_id, log_date, status_after, entered_by)
                       VALUES (%s, %s, %s, %s) RETURNING id""",
                    (work_order_id, r.parsed_date, r.status, self.imported_user_id),
                )
                log_entry_id = cur.fetchone()[0]
                for name in r.workers:
                    worker_id = self.get_worker_id(name)
                    if worker_id:
                        cur.execute(
                            "INSERT INTO staging.log_entry_workers (log_entry_id, worker_id) "
                            "VALUES (%s, %s) ON CONFLICT DO NOTHING",
                            (log_entry_id, worker_id),
                        )
                if r.unmatched_worker_text:
                    self.review.append(
                        ReviewItem(
                            "must_resolve", wo_no, r.excel_row,
                            f"Unrecognised worker name fragment '{r.unmatched_worker_text}' "
                            "in the By column — not attributed to anyone.",
                            r.by_raw,
                        )
                    )

        for r in rows:
            if r.parsed_date is None:
                self.review.append(
                    ReviewItem(
                        "must_resolve", wo_no, r.excel_row,
                        "Date cell is not a parseable date (e.g. 'ค้าง') — no log "
                        "entry created for this row.", str(r.date_raw),
                    )
                )
        return len(dated)


def read_rows(path: str) -> list[RawRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.worksheets[0]
    rows: list[RawRow] = []
    for excel_row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, min_col=1, max_col=8, values_only=True), start=2
    ):
        no, wo_no, date_raw, category, location, description, status, by = (
            list(row) + [None] * (8 - len(row))
        )[:8]

        wo_no_s = nfc(wo_no) if wo_no is not None else ""
        if not wo_no_s or wo_no_s in HEADER_MARKERS:
            continue  # §6 step 2: empty WO# or repeated header row
        if nfc(date_raw) == "Date" or nfc(status) == "Status":
            continue  # repeated header row detected via a different column

        rows.append(
            RawRow(
                excel_row=excel_row_idx,
                wo_no=wo_no_s,
                date_raw=date_raw,
                category_raw=nfc(category),
                location_raw=nfc(location),
                description=nfc(description),
                status_raw=nfc(status),
                by_raw=nfc(by) if by else "",
            )
        )
    return rows


def process(rows: list[RawRow], importer: Importer) -> dict:
    for r in rows:
        r.parsed_date = parse_date(r.date_raw)
        r.status = map_status(r.status_raw)
        r.workers, r.unmatched_worker_text = split_workers(r.by_raw)
        if r.status is None:
            importer.review.append(
                ReviewItem(
                    "must_resolve", r.wo_no, r.excel_row,
                    f"Unrecognised status value '{r.status_raw}' (expected 'เสร็จ' or 'ค้าง').",
                    r.status_raw,
                )
            )

    groups: dict[str, list[RawRow]] = {}
    for r in rows:
        groups.setdefault(r.wo_no, []).append(r)

    stats = {"numbers_in": len(groups), "rows_in": len(rows), "work_orders_out": 0, "log_entries_out": 0}

    for wo_no, group_rows in groups.items():
        usable = [r for r in group_rows if r.status is not None]
        if not usable:
            continue

        if wo_no in MANUAL_REVIEW_ONLY_WO:
            importer.review.append(
                ReviewItem(
                    "must_resolve", wo_no, [r.excel_row for r in group_rows],
                    "Genuine WO# collision touching a special (paid) category — "
                    "requires manual review before import, not auto-split (§6 step 5).",
                )
            )
            continue

        if wo_no == HARDCODED_MERGE_WO:
            importer.review.append(
                ReviewItem("fyi", wo_no, [r.excel_row for r in group_rows], HARDCODED_MERGE_REASON)
            )
            location_code = usable[0].location_raw
            n = importer.insert_work_order(wo_no, None, usable, HARDCODED_MERGE_CATEGORY, location_code)
            stats["work_orders_out"] += 1 if n else 0
            stats["log_entries_out"] += n
            continue

        # Apply the documented per-WO# category correction, if any, before
        # sub-grouping by (location, category) — so the correction is what
        # determines whether this group is actually a collision at all.
        if wo_no in CATEGORY_CORRECTIONS:
            from_cat, to_cat, reason = CATEGORY_CORRECTIONS[wo_no]
            for r in usable:
                if r.category_raw == from_cat:
                    importer.review.append(
                        ReviewItem("fyi", wo_no, r.excel_row, reason, r.category_raw)
                    )
                    r.category_raw = to_cat

        sub_groups: dict[tuple[str, str], list[RawRow]] = {}
        for r in usable:
            sub_groups.setdefault((r.location_raw, r.category_raw), []).append(r)

        if len(sub_groups) == 1:
            (location_code, category_th), sub_rows = next(iter(sub_groups.items()))
            n = importer.insert_work_order(wo_no, None, sub_rows, category_th, location_code)
            stats["work_orders_out"] += 1 if n else 0
            stats["log_entries_out"] += n
            continue

        # Genuine collision not on the manual-review list: auto-split, one
        # work order per distinct (location, category), first sub-group
        # keeps the original number, the rest get freshly generated ones,
        # and all of them carry the original in legacy_wo_no (§6 step 5).
        ordered = sorted(
            sub_groups.items(), key=lambda kv: min(r.excel_row for r in kv[1])
        )
        prefix = wo_no[:4]
        importer.review.append(
            ReviewItem(
                "fyi", wo_no, [r.excel_row for r in group_rows],
                f"Split automatically into {len(ordered)} work orders by (location, "
                "category) — different rooms/categories under one WO# (§6 step 5). Verify.",
            )
        )
        for i, ((location_code, category_th), sub_rows) in enumerate(ordered):
            new_wo_no = wo_no if i == 0 else importer.next_wo_no(prefix)
            n = importer.insert_work_order(new_wo_no, wo_no, sub_rows, category_th, location_code)
            stats["work_orders_out"] += 1 if n else 0
            stats["log_entries_out"] += n

    return stats


def write_review_file(review: list[ReviewItem], out_path: str) -> None:
    import openpyxl as _oxl

    wb = _oxl.Workbook()
    ws = wb.active
    ws.title = "review_needed"
    ws.append(["severity", "wo_no", "excel_row", "reason", "original_value"])
    for item in review:
        ws.append([item.severity, item.wo_no, str(item.excel_row), item.reason, item.original_value])
    wb.save(out_path)


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 scripts/import_excel.py <path-to-xlsx>", file=sys.stderr)
        sys.exit(1)
    xlsx_path = sys.argv[1]

    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        print("DATABASE_URL is not set", file=sys.stderr)
        sys.exit(1)

    rows = read_rows(xlsx_path)

    conn = psycopg2.connect(dsn)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            cur.execute("TRUNCATE staging.log_entry_workers, staging.wo_log_entries, staging.work_orders")
        importer = Importer(conn)
        stats = process(rows, importer)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    review_path = os.path.join(
        os.path.dirname(os.path.abspath(xlsx_path)) or ".", "review_needed.xlsx"
    )
    write_review_file(importer.review, review_path)

    must_resolve = sum(1 for i in importer.review if i.severity == "must_resolve")
    print("Import summary")
    print("--------------")
    print(f"Distinct WO# numbers read : {stats['numbers_in']}")
    print(f"Data rows read            : {stats['rows_in']}")
    print(f"Work orders staged        : {stats['work_orders_out']}")
    print(f"Log entries staged        : {stats['log_entries_out']}")
    print(f"Review items              : {len(importer.review)} ({must_resolve} must-resolve, "
          f"{len(importer.review) - must_resolve} fyi)")
    print(f"Review file               : {review_path}")
    print()
    print("Staged into the `staging` schema — nothing is in production yet.")
    print("Resolve review_needed.xlsx with your cousin, then promote with:")
    print("  psql \"$DATABASE_URL\" -f scripts/promote_staging.sql")


if __name__ == "__main__":
    main()
